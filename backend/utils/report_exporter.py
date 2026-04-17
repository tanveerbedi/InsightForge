# backend/utils/report_exporter.py
import csv
from io import BytesIO, StringIO
from typing import Any


def _get(outputs: dict, path: list[str], default: Any = None):
    current = outputs
    for part in path:
        if not isinstance(current, dict):
            return default
        current = current.get(part)
    return default if current is None else current


def _fmt(value):
    if isinstance(value, float):
        return f"{value:.4f}"
    if isinstance(value, (list, dict)):
        return str(value)[:300]
    return "" if value is None else str(value)


def export_to_pdf(pipeline_outputs: dict, run_id: str) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36)
    styles = getSampleStyleSheet()
    story = []

    ml = pipeline_outputs.get("ml_results", {})
    cleaning = pipeline_outputs.get("cleaning_results", {})
    eda = pipeline_outputs.get("eda_results", {})
    explain = pipeline_outputs.get("explainability_results", {})
    report = pipeline_outputs.get("report_results", {})

    story.append(Paragraph("InsightForge Analysis Report", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Run ID: {run_id}", styles["Normal"]))
    story.append(Paragraph(f"Date: {_fmt(pipeline_outputs.get('created_at'))}", styles["Normal"]))
    story.append(Spacer(1, 24))

    best_name = ml.get("best_model_name", "N/A")
    problem_type = ml.get("problem_type") or pipeline_outputs.get("problem_type", "unknown")
    best_metrics = ml.get("best_metrics", {})
    primary_metric = "f1_weighted" if problem_type == "classification" else "r2"
    primary_score = best_metrics.get(primary_metric)

    story.append(Paragraph("Executive Summary", styles["Heading2"]))
    summary = (
        f"Best model: {best_name}. Primary metric: {primary_metric}={_fmt(primary_score)}. "
        f"Problem type: {problem_type}. {report.get('executive_summary', '')}"
    )
    story.append(Paragraph(summary, styles["BodyText"]))
    story.append(Spacer(1, 14))

    story.append(Paragraph("Data Overview", styles["Heading2"]))
    shape = eda.get("shape") or cleaning.get("cleaned_shape") or []
    rows = shape[0] if len(shape) > 0 else "N/A"
    cols = shape[1] if len(shape) > 1 else "N/A"
    missing_counts = eda.get("missing_per_column", {})
    total_missing = sum(v for v in missing_counts.values() if isinstance(v, (int, float)))
    total_cells = rows * cols if isinstance(rows, int) and isinstance(cols, int) else 0
    missing_pct = (total_missing / total_cells * 100) if total_cells else 0
    columns = ", ".join(list(eda.get("dtypes", {}).keys())[:40])
    story.append(Paragraph(f"Rows: {rows}; Columns: {cols}; Missing values: {missing_pct:.2f}%", styles["BodyText"]))
    story.append(Paragraph(f"Columns: {columns}", styles["BodyText"]))
    story.append(Spacer(1, 14))

    story.append(Paragraph("Model Comparison", styles["Heading2"]))
    model_rows = [["Rank", "Model", primary_metric, "F1", "R2", "Train Time"]]
    for model in ml.get("all_models", []):
        metrics = model.get("tuned_metrics") or model.get("metrics") or {}
        model_rows.append(
            [
                _fmt(model.get("rank")),
                model.get("display_name") or model.get("name", ""),
                _fmt(metrics.get(primary_metric)),
                _fmt(metrics.get("f1_weighted")),
                _fmt(metrics.get("r2")),
                _fmt(model.get("training_time_sec")),
            ]
        )
    story.append(_styled_table(model_rows))
    story.append(Spacer(1, 14))

    story.append(Paragraph("Best Model", styles["Heading2"]))
    story.append(Paragraph(ml.get("why_best", "No model conclusion available."), styles["BodyText"]))
    params_rows = [["Parameter", "Value"]]
    for key, value in (ml.get("best_params") or {}).items():
        params_rows.append([str(key), _fmt(value)])
    story.append(_styled_table(params_rows))
    story.append(Spacer(1, 14))

    story.append(Paragraph("Top 10 SHAP Features", styles["Heading2"]))
    shap_rows = [["Feature", "Mean Absolute SHAP"]]
    for item in (explain.get("global_importance") or [])[:10]:
        shap_rows.append([item.get("feature", ""), _fmt(item.get("mean_abs_shap"))])
    story.append(_styled_table(shap_rows))
    story.append(Spacer(1, 14))

    story.append(Paragraph("Recommendations", styles["Heading2"]))
    for rec in (report.get("recommendations") or [])[:3]:
        story.append(Paragraph(f"- {rec}", styles["BodyText"]))

    doc.build(story)
    return buffer.getvalue()


def _styled_table(rows):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    table = Table(rows, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366f1")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def export_to_excel(pipeline_outputs: dict, run_id: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    ml = pipeline_outputs.get("ml_results", {})
    cleaning = pipeline_outputs.get("cleaning_results", {})
    eda = pipeline_outputs.get("eda_results", {})
    explain = pipeline_outputs.get("explainability_results", {})

    overview_rows = [
        ("Run ID", run_id),
        ("Dataset", pipeline_outputs.get("dataset_name", "")),
        ("Goal", pipeline_outputs.get("goal", "")),
        ("Problem Type", ml.get("problem_type", "")),
        ("Best Model", ml.get("best_model_name", "")),
        ("Best Metrics", str(ml.get("best_metrics", {}))),
        ("Data Quality Score", cleaning.get("data_quality_score", "")),
        ("Rows", (eda.get("shape") or [""])[0]),
        ("Columns", (eda.get("shape") or ["", ""])[1]),
    ]
    _write_rows(ws, ["Metric", "Value"], overview_rows)

    ws_models = wb.create_sheet("All Models")
    model_headers = ["rank", "name", "primary_metric", "f1_weighted", "r2", "training_time_sec", "tuned", "error"]
    model_rows = []
    problem_type = ml.get("problem_type", "classification")
    primary = "f1_weighted" if problem_type == "classification" else "r2"
    for model in ml.get("all_models", []):
        metrics = model.get("tuned_metrics") or model.get("metrics") or {}
        model_rows.append(
            [
                model.get("rank"),
                model.get("name"),
                metrics.get(primary),
                metrics.get("f1_weighted"),
                metrics.get("r2"),
                model.get("training_time_sec"),
                model.get("tuned"),
                model.get("error"),
            ]
        )
    _write_rows(ws_models, model_headers, model_rows)

    ws_params = wb.create_sheet("Best Model Params")
    _write_rows(ws_params, ["Parameter", "Value"], list((ml.get("best_params") or {}).items()))

    ws_shap = wb.create_sheet("SHAP Features")
    shap_rows = [
        [item.get("feature"), item.get("mean_abs_shap")]
        for item in (explain.get("global_importance") or [])
    ]
    _write_rows(ws_shap, ["Feature", "Mean Absolute SHAP"], shap_rows)

    ws_cleaning = wb.create_sheet("Cleaning Log")
    _write_rows(ws_cleaning, ["Action"], [[item] for item in cleaning.get("cleaning_log", [])])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_rows(ws, headers, rows):
    from openpyxl.styles import Font, PatternFill

    ws.append(headers)
    fill = PatternFill("solid", fgColor="6366F1")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for row in rows:
        ws.append(list(row))
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 60)


def export_summary_csv(pipeline_outputs: dict) -> str:
    output = StringIO()
    fields = [
        "model_name",
        "rank",
        "primary_metric",
        "f1",
        "training_time_sec",
        "tuned",
        "best_params_summary",
    ]
    writer = csv.DictWriter(output, fieldnames=fields)
    writer.writeheader()
    ml = pipeline_outputs.get("ml_results", {})
    problem_type = ml.get("problem_type", "classification")
    primary = "f1_weighted" if problem_type == "classification" else "r2"
    for model in ml.get("all_models", []):
        metrics = model.get("tuned_metrics") or model.get("metrics") or {}
        writer.writerow(
            {
                "model_name": model.get("name"),
                "rank": model.get("rank"),
                "primary_metric": metrics.get(primary),
                "f1": metrics.get("f1_weighted"),
                "training_time_sec": model.get("training_time_sec"),
                "tuned": model.get("tuned"),
                "best_params_summary": str(model.get("tuned_params") or {})[:250],
            }
        )
    return output.getvalue()
